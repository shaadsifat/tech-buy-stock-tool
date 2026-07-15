import io
from datetime import datetime

from openpyxl import Workbook

from app import models

HEADERS = [
    "Product Name",
    "Product Category",
    "Tech Buy Link",
    "Other website link",
    "Stock Status (Tech Buy)",
    "Stock Status (Other)",
    "Regular Price",
    "Sale Price",
    "Regular Price (Other)",
    "Sale Price (Other)",
    "Need Action",
    "Updated",
    "Fetched Status",
]


def build_workbook(product_ids=None):
    rows = models.get_export_data_for(product_ids) if product_ids else models.get_export_data()

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Update"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    for row in rows:
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=row["name"])
        ws.cell(row=r, column=2, value=row["category"])

        techbuy_cell = ws.cell(row=r, column=3, value="Tech Buy Link")
        techbuy_cell.hyperlink = row["techbuy_link"]
        techbuy_cell.style = "Hyperlink"

        other_cell = ws.cell(row=r, column=4, value="Other website link")
        other_cell.hyperlink = row["other_link"]
        other_cell.style = "Hyperlink"

        ws.cell(row=r, column=5, value=row["techbuy_stock"])
        ws.cell(row=r, column=6, value=row["other_stock"])
        ws.cell(row=r, column=7, value=row["techbuy_regular"])
        ws.cell(row=r, column=8, value=row["techbuy_sale"])
        ws.cell(row=r, column=9, value=row["other_regular"])
        ws.cell(row=r, column=10, value=row["other_sale"])
        ws.cell(row=r, column=11, value=row["need_action"])
        ws.cell(row=r, column=12, value=row["updated"])
        ws.cell(row=r, column=13, value=row["fetched_status"])

    widths = [26, 18, 15, 20, 22, 20, 14, 12, 20, 18, 12, 10, 14]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_filename():
    now = datetime.now()
    return f"Tech_Buy_Stock_Update_{now.strftime('%d%b_%Y_%I%M%p').lower()}.xlsx"
